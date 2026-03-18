import io
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .settings import UNNECESSARY_BRANDS
from .base_processing_files import BaseProcessingFiles


class GoodsMovementReport(BaseProcessingFiles):
    def __init__(self, file_bytes: bytes, file_name: str):
        super().__init__(file_bytes, file_name)
        self._processed_data = None

    @property
    def get_stream(self):
        """Получение потока"""
        return self._get_processed()

    @property
    def get_file_name(self):
        """Получение имени файла"""
        return self.file_name

    def _get_processed(self):
        if self._processed_data is None:
            self._processed_data = self.process()
        return self._processed_data

    def unnecessary_brand_del(self, row: dict) -> None:
        brand = row.get('Производитель', '').lower()
        if brand in UNNECESSARY_BRANDS:
            print(f'Бренд {brand} не нужен')
            row['К перемещению К'] = ''

    def openpyxl_open_file(self, stream: io.BytesIO):
        """Открытие файла с помощью openpyxl для сохранения в тот же файл"""
        wb = load_workbook(stream)
        ws = wb.active
        return wb, ws

    def openpyxl_save_file(self, workbook: Workbook):
        """Создание потока и сохранение файла с помощью openpyxl."""
        output_stream = io.BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream

    def openpyxl_update_file(self, data: dict, worksheet: Worksheet, df: pd.DataFrame):
        """Обновление файла"""
        headers_index = self.get_headers_index(df)
        for row_num, value_row in data.items():
            cell = worksheet.cell(row=row_num, column=headers_index)
            cell.value = value_row

        return worksheet

    def pandas_open_file(self, stream: io.BytesIO):
        """Открытие файла с помощью pandas"""
        return pd.read_excel(stream)

    def adding_the_amount_of_movement(self, row: dict) -> None:
        """Добавление кол-ва перемещения в столбец К перемещению К"""
        if row['Кол-во к перем.'] != '':
            row['К перемещению К'] = row['Кол-во к перем.']

    def multiplicity_check(self, row: dict) -> None:
        """Проверка кратности"""
        amout = row.get('К перемещению К', 0)
        mult = row.get('Кратность продажи', 1)

        if mult != 1 and amout:
            remainder = amout % mult
            if remainder != 0 and amout > mult:
                print(f'Остаток не кратен кратности, остаток {remainder}')
                row['К перемещению К'] = amout - remainder
            elif remainder != 0 and amout < mult:
                row['К перемещению К'] = mult

    def stock_check(self, row: dict) -> None:
        """Проверка остатка"""
        stock = row.get('Остаток отпр', 0)
        mult = row.get('Кратность продажи', 1)
        available_balance = row.get('Дост Ост Отпр', 1)

        if stock < mult or available_balance < mult:
            print(f'Остаток меньше кратности')
            row['К перемещению К'] = ''

        elif stock < mult * 2:
            print(f'Остаток меньше кратности * 2')
            row['К перемещению К'] = ''

    def validate_inventory(self, rows: list) -> dict:
        """Проверка данных отчёта"""
        update_rows = {}
        for i, row in enumerate(rows, start=2):
            if row['Кол-во к перем.'] != '':
                self.adding_the_amount_of_movement(row)
                self.multiplicity_check(row)
                self.stock_check(row)
                self.unnecessary_brand_del(row)
            update_rows[i] = row['К перемещению К']

        return update_rows

    def get_datafile(self, df: pd.DataFrame) -> dict:
        """Получение данных из файла"""
        rows = df.to_dict(orient='records')
        result = self.validate_inventory(rows)
        return result

    def del_nan(self, df: pd.DataFrame) -> pd.DataFrame:
        """Удаление NaN"""
        for col in df.columns:
            if col == 'Кратность продажи':
                df[col] = df[col].fillna(1)
            else:
                df[col] = df[col].fillna('')
        return df

    @staticmethod
    def get_headers_index(df: pd.DataFrame):
        return df.columns.tolist().index('К перемещению К') + 1

    def process(self) -> bytes:
        """Чтение файла"""
        input_stream = io.BytesIO(self.file_bytes)
        df = self.pandas_open_file(input_stream)
        input_stream.seek(0)
        workbook, worksheet = self.openpyxl_open_file(input_stream)
        input_stream.seek(0)
        df = self.del_nan(df)
        data = self.get_datafile(df)
        self.openpyxl_update_file(data, worksheet, df)
        output_stream = self.openpyxl_save_file(workbook)

        return output_stream.getvalue()
