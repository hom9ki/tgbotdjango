import io

from openpyxl import load_workbook


def remove_merge(file_bytes: bytes, file_name: str):
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    merges = ws.merged_cells

    for merge in list(merges):
        ws.unmerge_cells(range_string=str(merge))

    max_col = ws.max_column

    for col in range(1, max_col + 1):

        value1 = ws.cell(row=1, column=col).value or ''
        value2 = ws.cell(row=2, column=col).value or ''
        value3 = ws.cell(row=3, column=col).value or ''


        result = f'{value1} {value2} {value3}'.strip()

        ws.cell(row=1, column=col).value = ''
        ws.cell(row=2, column=col).value = result

    ws.delete_rows(3)
    output_stream = io.BytesIO()
    wb.save(output_stream)
    wb.close()

    output_stream.seek(0)
    return output_stream.getvalue()
