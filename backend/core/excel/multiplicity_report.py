import io
from openpyxl import load_workbook
from tqdm import tqdm
import pandas as pd
from .settings import key_words, anti_key_words, multiplicity


def normalize_text(text):
    return str(text).lower().strip() if text else ''


def miltiplicity_processing_excel(file_bytes):
    input_stream = io.BytesIO(file_bytes)
    workbook = load_workbook(input_stream, data_only=True)
    sheet = workbook.active

    data = []
    header = [cell.value for cell in sheet[1]]
    print(f'Header: {header}')
    for row in tqdm(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True),
                    desc='Processing'):
        data.append(row)

    df = pd.DataFrame(data, columns=header, index=None)
    df = df.fillna('')
    records = df.to_dict(orient='records')

    for record in tqdm(records, desc='Анализ:'):
        product = normalize_text(record.get('Наименование', ''))
        number = normalize_text(record.get('Номер по каталогу', ''))

        has_kit = any(kw in product for kw in ['комплект', 'набор', 'упаковка'])

        record['Кратность'] = 1

        if has_kit:
            continue

        if number.endswith(('l', 'r')) and not number.endswith('lr'):
            continue
        elif number.endswith('lr'):
            record['Кратность'] = 2
            continue

        for group, keywords in key_words.items():
            if any(word in product for word in keywords):
                anti_key = anti_key_words.get(group, [])
                normalized_words = product.replace('(', '').replace(')', '').split(' ')
                if anti_key and any(word in normalized_words for word in anti_key):
                    record['Кратность'] = 1
                    break
                elif anti_key and not any(word in normalized_words for word in anti_key):
                    record['Кратность'] = multiplicity[group]
                    break
                record['Кратность'] = multiplicity[group]

    for ind, record in enumerate(records, start=2):
        sheet[f'D{ind}'] = record.get('Кратность', 1)

    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    return output_stream.read()
